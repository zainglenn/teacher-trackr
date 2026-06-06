from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Palette ──────────────────────────────────────────────────────────────────
NAVY       = "1A1A2E"
WHITE      = "FFFFFF"
GOLD       = "C9A02C"
LIGHT_GREY = "F4F4F8"
MID_GREY   = "E2E2E8"
DARK_GREY  = "888899"

P0_COLOR   = "FEE2E2"  # red-100
P1_COLOR   = "FEF3C7"  # amber-100
P2_COLOR   = "D1FAE5"  # green-100
P0_TEXT    = "B91C1C"
P1_TEXT    = "B45309"
P2_TEXT    = "065F46"

SUITE_FILLS = [
    "DBEAFE", "EDE9FE", "D1FAE5", "FEF3C7", "FCE7F3",
    "E0F2FE", "F0FDF4", "FFF7ED", "F5F3FF", "ECFDF5",
]

def col_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color=MID_GREY)
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=10, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=9, bold=False, color="1A1A2E"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="top", wrap_text=True)


# ── Test data ─────────────────────────────────────────────────────────────────

COLUMNS = ["ID", "Suite", "Test Case", "Preconditions", "Steps", "Expected Result", "Priority", "Status", "Notes"]
COL_WIDTHS = [10, 28, 38, 32, 52, 40, 10, 12, 28]

# ── ROLE 1: Platform Admin ────────────────────────────────────────────────────

PLATFORM_CASES = [
    # Suite, ID, Name, Preconditions, Steps, Expected, Priority
    ("Authentication", "PA-AUTH-01", "Login as platform admin",
     "Account zain.platform exists",
     "1. Go to app URL\n2. Enter username: zain.platform\n3. Enter password\n4. Click Sign in",
     "Lands on Schools page. Sidebar shows Schools and Curricula nav items only.",
     "P0"),
    ("Authentication", "PA-AUTH-02", "Platform admin cannot access school admin routes",
     "Logged in as zain.platform",
     "1. Call POST /api/admin/create-standard-set with platform admin token",
     "Returns 403 Forbidden",
     "P1"),
    ("Authentication", "PA-AUTH-03", "Platform admin cannot access teacher routes",
     "Logged in as zain.platform",
     "1. Call GET /api/admin/list-users with platform admin token",
     "Returns 403 Forbidden",
     "P1"),
    ("Authentication", "PA-AUTH-04", "Sign out",
     "Logged in as zain.platform",
     "1. Click Sign out in sidebar footer",
     "Returns to login screen. Session cleared.",
     "P1"),

    ("Schools", "PA-SCH-01", "Schools page loads",
     "Logged in as zain.platform",
     "1. Click Schools in sidebar",
     "Page title = 'Schools'. Table shows all schools with Name, City, Curriculum, Users, Created, Status columns. No tab switcher visible.",
     "P0"),
    ("Schools", "PA-SCH-02", "Search schools by name",
     "Multiple schools exist",
     "1. Type school name in search box",
     "Table filters live. Non-matching schools hidden.",
     "P1"),
    ("Schools", "PA-SCH-03", "Search schools by city",
     "Schools with different cities exist",
     "1. Type a city name in search box",
     "Table filters to matching city only.",
     "P1"),
    ("Schools", "PA-SCH-04", "Create new school",
     "Logged in as zain.platform",
     "1. Click New School\n2. Fill name, city, country, curriculum type\n3. Click Create",
     "School appears in list. Active status badge shown.",
     "P1"),
    ("Schools", "PA-SCH-05", "View school detail sheet",
     "At least one school exists",
     "1. Click any school row",
     "Detail sheet slides in showing school info and user count.",
     "P1"),
    ("Schools", "PA-SCH-06", "Suspend a school — confirmation required",
     "Active school exists",
     "1. Click ⋯ on a school → Suspend\n2. Observe modal\n3. Type wrong name\n4. Try to confirm",
     "Modal opens with school name. Confirm button disabled until exact name typed.",
     "P0"),
    ("Schools", "PA-SCH-07", "Suspend school — confirm",
     "Active school exists",
     "1. Click ⋯ → Suspend\n2. Type exact school name\n3. Click Suspend",
     "School status changes to Suspended. Users at that school see suspended message.",
     "P0"),
    ("Schools", "PA-SCH-08", "Unsuspend a school",
     "Suspended school exists",
     "1. Click ⋯ → Unsuspend",
     "School status changes to Active immediately.",
     "P1"),

    ("Curricula", "PA-CUR-01", "Navigate to Curricula page",
     "Logged in as zain.platform",
     "1. Click Curricula in sidebar",
     "Full page navigation. Title = 'Curricula'. Separate from Schools page — no shared state.",
     "P0"),
    ("Curricula", "PA-CUR-02", "Curricula list loads",
     "Platform standard sets exist in DB",
     "1. On Curricula page",
     "All platform standard sets shown. Each row shows name and subject/grade labels.",
     "P0"),
    ("Curricula", "PA-CUR-03", "Empty state",
     "No standard sets in DB",
     "1. Navigate to Curricula page",
     "BookOpen icon + 'No curricula yet' message with New Curriculum button.",
     "P2"),
    ("Curricula", "PA-CUR-04", "Create new curriculum — validation",
     "On Curricula page",
     "1. Click New Curriculum\n2. Leave Name empty\n3. Attempt to click Create",
     "Create button disabled while name is empty.",
     "P0"),
    ("Curricula", "PA-CUR-05", "Create new curriculum",
     "On Curricula page",
     "1. Click New Curriculum\n2. Enter Name: 'NYSED Grade 9 ELA'\n3. Subject label: 'English'\n4. Grade label: 'Grade 9'\n5. Click Create",
     "New row appears in list with correct name and labels. Form closes.",
     "P0"),
    ("Curricula", "PA-CUR-06", "Expand curriculum row",
     "At least one curriculum with standards",
     "1. Click row or chevron to expand",
     "Standards table appears below row. Table has Code, Strand, Description, delete columns.",
     "P1"),
    ("Curricula", "PA-CUR-07", "Add standard — Strand is shadcn Select",
     "Curriculum row expanded",
     "1. Click Add Standard\n2. Observe Strand field",
     "Strand field uses shadcn Select component (not native <select>). Dropdown shows RL, RI, W, SL, L.",
     "P0"),
    ("Curricula", "PA-CUR-08", "Add standard to curriculum",
     "Curriculum expanded, Add Standard form open",
     "1. Enter Code: 'RL.9.1'\n2. Select Strand: RL\n3. Enter Description\n4. Click Add",
     "Standard appears in table sorted by code. Form resets.",
     "P0"),
    ("Curricula", "PA-CUR-09", "Delete standard from curriculum",
     "Curriculum has at least one standard",
     "1. Click trash icon on a standard row",
     "Standard removed immediately. No confirmation needed.",
     "P1"),
    ("Curricula", "PA-CUR-10", "Delete curriculum — modal appears",
     "At least one curriculum exists",
     "1. Click trash icon on curriculum row header",
     "Confirmation modal opens. Shows curriculum name, warns about cascading delete of standards and school assignments.",
     "P0"),
    ("Curricula", "PA-CUR-11", "Delete curriculum — cancel",
     "Delete modal open",
     "1. Click Cancel in modal",
     "Modal closes. Curriculum remains in list unchanged.",
     "P0"),
    ("Curricula", "PA-CUR-12", "Delete curriculum — confirm",
     "Delete modal open",
     "1. Click Delete in modal",
     "Curriculum removed from list. Any schools assigned to it lose the assignment.",
     "P0"),
    ("Curricula", "PA-CUR-13", "Collapsed row does not show standards count",
     "Curriculum exists with standards",
     "1. View collapsed curriculum row",
     "Row shows name and labels only. Standards are visible after expansion.",
     "P2"),

    ("API Security", "PA-SEC-01", "Only platform_admin can create curricula",
     "School admin token available",
     "1. POST /api/platform/curricula with school admin token",
     "Returns 403 Forbidden",
     "P1"),
    ("API Security", "PA-SEC-02", "Only platform_admin can delete curricula",
     "Teacher token available",
     "1. DELETE /api/platform/curricula with teacher token",
     "Returns 403 Forbidden",
     "P1"),
    ("API Security", "PA-SEC-03", "Only platform_admin can add standards",
     "School admin token available",
     "1. POST /api/platform/curricula/standards with school admin token",
     "Returns 403 Forbidden",
     "P1"),
]

# ── ROLE 2: School Admin ──────────────────────────────────────────────────────

ADMIN_CASES = [
    ("Authentication", "AD-AUTH-01", "Login as school admin",
     "Account zain.admin exists",
     "1. Enter username: zain.admin\n2. Enter password\n3. Click Sign in",
     "Lands on Manage Users page. Sidebar shows admin nav items (Manage Users, School Setup, Platform Settings, Curriculum Audit).",
     "P0"),
    ("Authentication", "AD-AUTH-02", "Admin cannot access platform routes",
     "Logged in as zain.admin",
     "1. POST /api/platform/curricula with admin token",
     "Returns 403 Forbidden",
     "P1"),
    ("Authentication", "AD-AUTH-03", "Admin cannot access other school data",
     "Two schools exist",
     "1. GET /api/admin/school-curricula with School A admin token",
     "Only returns School A's curriculum assignments. School B data not visible.",
     "P0"),

    ("Manage Users", "AD-USR-01", "View users list",
     "Logged in as zain.admin",
     "1. Click Manage Users",
     "Table shows all users with username, role badge, joined date.",
     "P1"),
    ("Manage Users", "AD-USR-02", "Add teacher user",
     "On Manage Users",
     "1. Click Add User\n2. Set role: Teacher\n3. Fill username and password\n4. Click Save",
     "New teacher appears in list with Teacher badge.",
     "P1"),
    ("Manage Users", "AD-USR-03", "Add HOD user",
     "On Manage Users",
     "1. Click Add User\n2. Set role: HOD\n3. Fill details\n4. Save",
     "New HOD appears with HOD badge.",
     "P1"),
    ("Manage Users", "AD-USR-04", "Filter users by role",
     "Multiple users with different roles",
     "1. Change role filter dropdown",
     "List filters to selected role only.",
     "P2"),

    ("School Setup — Subjects", "AD-SUB-01", "View subjects",
     "On School Setup",
     "1. Click School Setup → Subjects tab",
     "All school subjects listed with colour badges.",
     "P1"),
    ("School Setup — Subjects", "AD-SUB-02", "Add subject",
     "On Subjects tab",
     "1. Click + Add Subject\n2. Enter name\n3. Save",
     "Subject appears with auto-assigned colour slot.",
     "P1"),
    ("School Setup — Subjects", "AD-SUB-03", "Delete subject",
     "At least one subject exists",
     "1. Click trash on a subject",
     "Subject removed from list.",
     "P2"),

    ("School Setup — Grade Levels", "AD-GRD-01", "View grade levels",
     "On School Setup",
     "1. Click Grade Levels tab",
     "All grade levels listed in sort order.",
     "P1"),
    ("School Setup — Grade Levels", "AD-GRD-02", "Add grade level",
     "On Grade Levels tab",
     "1. Click + Add Grade Level\n2. Enter name: 'Grade 9'\n3. Save",
     "Grade 9 appears in list.",
     "P1"),

    ("School Setup — Standard Sets", "AD-STD-01", "Navigate to Standard Sets tab",
     "School has subjects and grade levels",
     "1. School Setup → Standard Sets tab",
     "Page shows Subject and Grade Level dropdowns. Description reads 'Select a subject and grade level, then choose the platform curriculum that applies.'",
     "P0"),
    ("School Setup — Standard Sets", "AD-STD-02", "Select subject and grade — no assignment",
     "No curriculum assigned for this combo",
     "1. Select a subject\n2. Select a grade level",
     "Shows 'No curriculum assigned' message. Search input and full platform library list visible.",
     "P0"),
    ("School Setup — Standard Sets", "AD-STD-03", "Library shows all platform curricula",
     "Multiple platform curricula exist",
     "1. Subject and grade selected, unassigned state",
     "All platform curricula shown regardless of subject/grade labels. Not pre-filtered.",
     "P0"),
    ("School Setup — Standard Sets", "AD-STD-04", "Search filters by name",
     "Library visible",
     "1. Type 'NYSED' in search input",
     "Only curricula with 'NYSED' in name shown.",
     "P0"),
    ("School Setup — Standard Sets", "AD-STD-05", "Search filters by subject label",
     "Library visible",
     "1. Type 'english' in search (case-insensitive)",
     "Only curricula with subject_label containing 'english' shown.",
     "P0"),
    ("School Setup — Standard Sets", "AD-STD-06", "Search filters by grade label",
     "Library visible",
     "1. Type 'grade 6' in search",
     "Only curricula with grade_label containing 'grade 6' shown.",
     "P0"),
    ("School Setup — Standard Sets", "AD-STD-07", "Search no results",
     "Library visible",
     "1. Type a string matching nothing",
     "Shows 'No curricula match \"...\"' message.",
     "P1"),
    ("School Setup — Standard Sets", "AD-STD-08", "Clear search restores list",
     "Search has filtered results",
     "1. Clear the search input",
     "Full platform library restored.",
     "P1"),
    ("School Setup — Standard Sets", "AD-STD-09", "Assign curriculum",
     "Library visible",
     "1. Click Assign on a curriculum",
     "View switches to assigned state: curriculum name, standard count, Remove button. Standards table renders read-only.",
     "P0"),
    ("School Setup — Standard Sets", "AD-STD-10", "Assigned standards table is read-only",
     "Curriculum assigned",
     "1. View assigned curriculum",
     "Standards table shows Code, Strand, Description. No delete or edit controls.",
     "P1"),
    ("School Setup — Standard Sets", "AD-STD-11", "Remove assignment",
     "Curriculum assigned",
     "1. Click Remove",
     "Returns to unassigned state with library list visible.",
     "P0"),
    ("School Setup — Standard Sets", "AD-STD-12", "Separate combos are independent",
     "School has multiple subjects/grades",
     "1. Assign curriculum to English / Grade 6\n2. Switch to Maths / Grade 6",
     "Maths / Grade 6 shows as unassigned independently.",
     "P1"),
    ("School Setup — Standard Sets", "AD-STD-13", "create-standard-set API returns 410",
     "School admin token",
     "1. POST /api/admin/create-standard-set",
     "Returns 410 Gone with message about platform admin route.",
     "P1"),

    ("School Setup — Class Assignments", "AD-CLS-01", "View class assignments",
     "On School Setup",
     "1. Class Assignments tab\n2. Select subject + grade",
     "Shows assigned teachers with Lead indicator.",
     "P1"),
    ("School Setup — Class Assignments", "AD-CLS-02", "Assign teacher to class",
     "Teachers exist, combo selected",
     "1. Select teacher from dropdown\n2. Click Assign",
     "Teacher appears in assignment list.",
     "P1"),
    ("School Setup — Class Assignments", "AD-CLS-03", "Toggle lead teacher",
     "Teacher assigned",
     "1. Click lead toggle on a teacher",
     "Lead indicator updates.",
     "P2"),
    ("School Setup — Class Assignments", "AD-CLS-04", "Remove class assignment",
     "Teacher assigned",
     "1. Click remove on teacher",
     "Teacher removed from assignment.",
     "P2"),
]

# ── ROLE 3: Teacher ───────────────────────────────────────────────────────────

TEACHER_CASES = [
    ("Authentication", "TC-AUTH-01", "Login as teacher",
     "Account jade.teacher exists",
     "1. Enter username: jade.teacher\n2. Enter password\n3. Sign in",
     "Lands on My Units page. Teacher sidebar visible.",
     "P0"),
    ("Authentication", "TC-AUTH-02", "Teacher cannot access admin routes",
     "Logged in as jade.teacher",
     "1. POST /api/admin/create-standard-set with teacher token",
     "Returns 403 Forbidden",
     "P1"),
    ("Authentication", "TC-AUTH-03", "Teacher cannot access platform routes",
     "Logged in as jade.teacher",
     "1. GET /api/platform/list-schools with teacher token",
     "Returns 403 Forbidden",
     "P1"),

    ("Standards Loading", "TC-STD-01", "Standards load via school_curricula",
     "School has curriculum assigned for teacher's subject/grade",
     "1. Log in as teacher\n2. Navigate to Standards Coverage",
     "Standards visible and grouped by strand. Loaded from school_curricula join, not old standard_sets.school_id.",
     "P0"),
    ("Standards Loading", "TC-STD-02", "No standards if no curriculum assigned",
     "School has NO curriculum assigned for this combo",
     "1. Log in as teacher with no assigned curriculum",
     "Coverage view shows empty state. No standards listed.",
     "P1"),
    ("Standards Loading", "TC-STD-03", "Standards scoped to school only",
     "Two schools with different curricula",
     "1. Log in as teacher from School A",
     "Only standards from School A's assigned curricula visible. School B standards not shown.",
     "P0"),

    ("My Units", "TC-UNT-01", "My Units page loads",
     "Teacher has assigned units",
     "1. Click My Units",
     "List of assigned unit plans visible.",
     "P1"),
    ("My Units", "TC-UNT-02", "Open unit plan",
     "At least one unit assigned",
     "1. Click a unit title",
     "Unit plan view opens with standards table, essential question, assessments.",
     "P1"),

    ("Long Term Plans", "TC-LTP-01", "View master plans",
     "LTP exists for teacher",
     "1. Click Master Plans",
     "Teacher's LTPs listed with status badges.",
     "P1"),
    ("Long Term Plans", "TC-LTP-02", "Create new unit",
     "LTP exists",
     "1. Open LTP\n2. Click Add Unit\n3. Fill title, big idea, duration\n4. Create",
     "Unit created, navigates to UnitPlanView.",
     "P1"),
    ("Long Term Plans", "TC-LTP-03", "Map standards to unit",
     "Unit open in UnitPlanView, standards loaded",
     "1. Open Map Standards section\n2. Select standards\n3. Save",
     "Standards appear in unit standards table sorted by strand.",
     "P0"),
    ("Long Term Plans", "TC-LTP-04", "AI suggest standards",
     "Unit with essential question, DeepSeek key set",
     "1. Open Map Standards\n2. Click AI Suggest",
     "3–6 standards suggested based on unit theme.",
     "P2"),
    ("Long Term Plans", "TC-LTP-05", "Submit unit for review",
     "Unit has standards mapped",
     "1. Open unit\n2. Click Submit for Review",
     "Unit status changes to Submitted. HOD receives notification.",
     "P0"),
    ("Long Term Plans", "TC-LTP-06", "Edit unit in Draft state",
     "Unit in Draft status",
     "1. Open unit\n2. Edit title\n3. Save Changes",
     "Changes persisted.",
     "P1"),
    ("Long Term Plans", "TC-LTP-07", "Cannot edit Approved unit",
     "Unit in Approved status",
     "1. Open approved unit",
     "Fields read-only. Save Changes button not available.",
     "P1"),

    ("Standards Coverage", "TC-COV-01", "Coverage view loads",
     "Standards assigned via school_curricula",
     "1. Click Standards Coverage",
     "Standards grouped by strand with coverage progress bars.",
     "P1"),
    ("Standards Coverage", "TC-COV-02", "Coverage reflects mapped standards",
     "Some standards mapped in LTP units",
     "1. Check coverage % vs units' mapped standards",
     "Coverage percentage matches standards in approved/submitted units.",
     "P1"),

    ("Student Progress", "TC-STU-01", "Student progress view loads",
     "Students enrolled",
     "1. Click Student Progress",
     "Students listed with attainment levels per standard.",
     "P1"),

    ("HOD — Review", "TC-HOD-01", "HOD login — lands on Dashboard",
     "Account hod.test exists",
     "1. Login as hod.test",
     "HOD Dashboard loads with coverage heatmap by teacher × strand.",
     "P0"),
    ("HOD — Review", "TC-HOD-02", "HOD coverage heatmap loads",
     "Teachers have mapped standards via school_curricula",
     "1. View Dashboard",
     "Heatmap shows teachers as rows, strands as columns, % values in cells.",
     "P0"),
    ("HOD — Review", "TC-HOD-03", "Plan Reviews — submitted units appear",
     "Teacher has submitted a unit",
     "1. Click Plan Reviews",
     "Submitted units listed in FIFO order with teacher name and submission date.",
     "P0"),
    ("HOD — Review", "TC-HOD-04", "Approve a unit",
     "Unit in Submitted status",
     "1. Click Approve on a unit",
     "Unit status changes to Approved. Teacher can no longer edit.",
     "P0"),
    ("HOD — Review", "TC-HOD-05", "Request revision",
     "Unit in Submitted status",
     "1. Click Revise\n2. Enter feedback text\n3. Submit",
     "Unit status changes to Revision. Feedback visible to teacher.",
     "P0"),
    ("HOD — Review", "TC-HOD-06", "Reject a unit",
     "Unit in Submitted status",
     "1. Click Reject\n2. Enter rejection reason\n3. Confirm",
     "Unit status changes to Rejected. Teacher must contact HOD to reopen.",
     "P1"),
    ("HOD — Review", "TC-HOD-07", "HOD cannot review without feedback on revision",
     "Revision modal open",
     "1. Open Revise modal\n2. Leave feedback blank\n3. Attempt submit",
     "Submit button disabled until feedback entered.",
     "P1"),
    ("HOD — Review", "TC-HOD-08", "Standards gap report shows per plan",
     "Plan with missing standards",
     "1. On Plan Reviews, expand a plan",
     "Gap report shows unmapped standards by strand.",
     "P1"),

    ("Regression", "TC-REG-01", "Existing approved units retain standards",
     "Units approved before architecture change",
     "1. Open any pre-existing approved unit",
     "Standards still display correctly. No data loss from migration.",
     "P0"),
    ("Regression", "TC-REG-02", "Delivery grid loads",
     "HOD logged in",
     "1. Click Delivery Grid",
     "Grid shows units across teachers and time. No errors.",
     "P1"),
    ("Regression", "TC-REG-03", "Notifications load",
     "Teacher logged in",
     "1. Check notification bell",
     "Notifications appear (overdue, pending review, etc.).",
     "P1"),
]


# ── Build workbook ─────────────────────────────────────────────────────────────

wb = Workbook()

def make_sheet(wb, title, role_color, cases, role_label):
    ws = wb.create_sheet(title=title)

    # ── Title row ──────────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = f"Test Plan — {role_label}  |  Curriculum Tracker 2026–2027"
    c.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c.fill = col_fill(NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 28

    # ── Sub-header row ─────────────────────────────────────────────────────────
    ws.merge_cells("A2:I2")
    c2 = ws["A2"]
    c2.value = f"Role: {role_label}   |   Total cases: {len(cases)}   |   Generated: 2026-06-04"
    c2.font = Font(name="Calibri", size=9, color=DARK_GREY)
    c2.fill = col_fill(LIGHT_GREY)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Column headers ─────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 22
    for i, (col_name, width) in enumerate(zip(COLUMNS, COL_WIDTHS), start=1):
        cell = ws.cell(row=3, column=i, value=col_name)
        cell.font = header_font(size=9)
        cell.fill = col_fill(role_color)
        cell.alignment = center()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── Data rows ──────────────────────────────────────────────────────────────
    suite_color_map = {}
    color_idx = 0
    suite_start_rows = {}

    for row_idx, (suite, tc_id, name, pre, steps, expected, priority) in enumerate(cases, start=4):
        # Assign suite colour
        if suite not in suite_color_map:
            suite_color_map[suite] = SUITE_FILLS[color_idx % len(SUITE_FILLS)]
            color_idx += 1
            suite_start_rows[suite] = row_idx

        s_color = suite_color_map[suite]

        # Priority badge colours
        if priority == "P0":
            p_fill, p_font = col_fill(P0_COLOR), Font(name="Calibri", size=9, bold=True, color=P0_TEXT)
        elif priority == "P1":
            p_fill, p_font = col_fill(P1_COLOR), Font(name="Calibri", size=9, bold=True, color=P1_TEXT)
        else:
            p_fill, p_font = col_fill(P2_COLOR), Font(name="Calibri", size=9, bold=True, color=P2_TEXT)

        row_data = [tc_id, suite, name, pre, steps, expected, priority, "", ""]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border()

            if col_idx == 2:  # Suite
                cell.fill = col_fill(s_color)
                cell.font = body_font(bold=True)
                cell.alignment = left()
            elif col_idx == 7:  # Priority
                cell.fill = p_fill
                cell.font = p_font
                cell.alignment = center()
            elif col_idx == 8:  # Status (dropdown-style)
                cell.fill = col_fill(LIGHT_GREY)
                cell.font = body_font(color=DARK_GREY)
                cell.alignment = center()
                cell.value = "⬜ Untested"
            elif col_idx in (5, 6):  # Steps, Expected — taller
                cell.alignment = left()
                cell.font = body_font(size=8.5)
            else:
                cell.alignment = left()
                cell.font = body_font()

        ws.row_dimensions[row_idx].height = max(
            len(steps.split("\n")) * 13,
            13
        )

    # ── Suite group headers ────────────────────────────────────────────────────
    # Add subtle suite header rows (insert logic is complex so we style in-place)
    # Instead, bold the first row of each new suite in col B
    current_suite = None
    for row_idx, (suite, *_) in enumerate(cases, start=4):
        if suite != current_suite:
            current_suite = suite
            cell = ws.cell(row=row_idx, column=2)
            cell.font = Font(name="Calibri", size=9, bold=True, color="1A1A2E")

    # ── Freeze top rows ────────────────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Auto-filter ────────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A3:I{3 + len(cases)}"

    return ws


# ── Summary sheet ──────────────────────────────────────────────────────────────

def make_summary(wb, all_roles):
    ws = wb.create_sheet(title="Summary", index=0)

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Test Plan Summary — Curriculum Tracker 2026–2027"
    c.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    c.fill = col_fill(NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    c2 = ws["A2"]
    c2.value = "Generated: 2026-06-04   |   App: curriculum-tracker-five.vercel.app   |   Branch: main"
    c2.font = Font(name="Calibri", size=9, color=DARK_GREY)
    c2.fill = col_fill(LIGHT_GREY)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # Headers
    headers = ["Role", "Test Suite", "Total Cases", "P0", "P1", "P2"]
    widths =   [22,     30,           14,             8,    8,    8]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = header_font()
        cell.fill = col_fill(NAVY)
        cell.alignment = center()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 20

    row = 4
    grand_total = grand_p0 = grand_p1 = grand_p2 = 0

    for role_label, role_color, cases in all_roles:
        suites = {}
        for suite, _, _, _, _, _, priority in cases:
            if suite not in suites:
                suites[suite] = {"total": 0, "P0": 0, "P1": 0, "P2": 0}
            suites[suite]["total"] += 1
            suites[suite][priority] += 1

        first = True
        for suite_name, counts in suites.items():
            if first:
                ws.cell(row=row, column=1, value=role_label).font = Font(name="Calibri", size=9, bold=True)
                ws.cell(row=row, column=1).fill = col_fill(role_color)
                first = False
            else:
                ws.cell(row=row, column=1, value="").fill = col_fill(role_color)

            ws.cell(row=row, column=2, value=suite_name).font = body_font()
            ws.cell(row=row, column=3, value=counts["total"]).alignment = center()
            ws.cell(row=row, column=4, value=counts["P0"]).fill = col_fill(P0_COLOR)
            ws.cell(row=row, column=4).font = Font(name="Calibri", size=9, bold=bool(counts["P0"]), color=P0_TEXT)
            ws.cell(row=row, column=4).alignment = center()
            ws.cell(row=row, column=5, value=counts["P1"]).fill = col_fill(P1_COLOR)
            ws.cell(row=row, column=5).font = Font(name="Calibri", size=9, color=P1_TEXT)
            ws.cell(row=row, column=5).alignment = center()
            ws.cell(row=row, column=6, value=counts["P2"]).fill = col_fill(P2_COLOR)
            ws.cell(row=row, column=6).font = Font(name="Calibri", size=9, color=P2_TEXT)
            ws.cell(row=row, column=6).alignment = center()

            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border()
                ws.row_dimensions[row].height = 16

            grand_total += counts["total"]
            grand_p0 += counts["P0"]
            grand_p1 += counts["P1"]
            grand_p2 += counts["P2"]
            row += 1

    # Grand total row
    ws.row_dimensions[row].height = 20
    for col, val in enumerate([" TOTAL", "", grand_total, grand_p0, grand_p1, grand_p2], start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = col_fill(NAVY)
        cell.alignment = center()
        cell.border = thin_border()

    # Legend
    ws.row_dimensions[row + 2].height = 16
    ws.cell(row=row+2, column=1, value="Priority Key").font = Font(name="Calibri", size=9, bold=True)
    for r, (label, fc, tc) in enumerate([("P0 — Critical (test before any deploy)", P0_COLOR, P0_TEXT),
                                          ("P1 — Important (test in each cycle)", P1_COLOR, P1_TEXT),
                                          ("P2 — Nice to have (regression/edge cases)", P2_COLOR, P2_TEXT)], start=row+3):
        c = ws.cell(row=r, column=1, value=label)
        c.fill = col_fill(fc)
        c.font = Font(name="Calibri", size=9, color=tc)
        ws.merge_cells(f"A{r}:F{r}")
        ws.row_dimensions[r].height = 15

    ws.freeze_panes = "A4"
    return ws


# ── ADDITIONAL: Admin Delete Operations ──────────────────────────────────────

ADMIN_DELETE_CASES = [
    ("Delete Operations", "AD-DEL-01", "Delete a user",
     "At least one non-admin user exists",
     "1. Manage Users\n2. Click ⋯ on a user → Delete\n3. Confirm",
     "User removed from list. They can no longer log in.",
     "P1"),
    ("Delete Operations", "AD-DEL-02", "Cannot delete own account",
     "Logged in as zain.admin",
     "1. Find own username in users list\n2. Check options",
     "No delete option available for own account.",
     "P1"),
    ("Delete Operations", "AD-DEL-03", "Delete a grade level",
     "Grade level with no class assignments exists",
     "1. School Setup → Grade Levels\n2. Click trash on a grade level",
     "Grade level removed.",
     "P2"),
    ("Delete Operations", "AD-DEL-04", "Delete a subject",
     "Subject with no class assignments exists",
     "1. School Setup → Subjects\n2. Click trash on a subject",
     "Subject removed.",
     "P2"),
    ("Delete Operations", "AD-DEL-05", "Update user role",
     "Teacher user exists",
     "1. Manage Users → click edit on a teacher\n2. Change role to HOD\n3. Save",
     "User badge updates to HOD. User sees HOD views on next login.",
     "P1"),
    ("Delete Operations", "AD-DEL-06", "Curriculum Audit — view all plans",
     "Teachers have LTPs",
     "1. Click Curriculum Audit\n2. View Long Term Plans tab",
     "Read-only table of all school LTPs with title, teacher, year, status, unit count.",
     "P1"),
    ("Delete Operations", "AD-DEL-07", "Curriculum Audit — view unit plans",
     "LTPs with units exist",
     "1. Curriculum Audit → Unit Plans tab",
     "All unit plans listed across all teachers.",
     "P2"),
    ("Delete Operations", "AD-DEL-08", "Platform Settings loads",
     "Logged in as zain.admin",
     "1. Click Platform Settings",
     "Settings page renders without error.",
     "P2"),
]

# ── ADDITIONAL: HOD Leadership Suite ─────────────────────────────────────────

HOD_LEADERSHIP_CASES = [
    ("HOD Settings", "HL-SET-01", "HOD Settings page loads",
     "Logged in as hod.test",
     "1. Click HOD Settings",
     "Settings page renders. HOD can update profile/notification preferences.",
     "P1"),

    ("Analytics", "HL-ANA-01", "Department analytics page loads",
     "Logged in as hod.test",
     "1. Click Analytics in sidebar",
     "DepartmentAnalyticsView renders without errors.",
     "P1"),
    ("Analytics", "HL-ANA-02", "Analytics show strand performance data",
     "Teachers have coverage logs",
     "1. View Analytics page",
     "Charts/stats show performance by strand across department.",
     "P1"),
    ("Analytics", "HL-ANA-03", "Grade level filter on analytics",
     "Multiple grade levels with data",
     "1. Change grade level filter",
     "Data updates to reflect selected grade level.",
     "P2"),

    ("Coaching", "HL-COA-01", "Coaching page loads",
     "Logged in as hod.test",
     "1. Click Coaching in sidebar",
     "CoachingView renders without errors.",
     "P1"),
    ("Coaching", "HL-COA-02", "View coaching cycles",
     "Coaching cycles exist",
     "1. View Coaching page",
     "List of coaching cycles with teacher name, steps completed, status.",
     "P1"),
    ("Coaching", "HL-COA-03", "Create observation",
     "On Coaching page",
     "1. Click New Observation\n2. Select teacher, focus area, date\n3. Add notes\n4. Save",
     "Observation saved and appears in list.",
     "P1"),
    ("Coaching", "HL-COA-04", "Progress a coaching cycle step",
     "Active coaching cycle exists",
     "1. Open a cycle\n2. Mark a step complete",
     "Step marked, cycle progress updates.",
     "P2"),

    ("Department", "HL-DEP-01", "Department view loads",
     "Logged in as hod.test or teacher",
     "1. Click Department in sidebar",
     "Department view renders without errors.",
     "P1"),
    ("Department", "HL-DEP-02", "Department view shows teacher list",
     "Multiple teachers in school",
     "1. View Department page",
     "Teachers listed with their coverage stats.",
     "P1"),

    ("Initiatives", "HL-INI-01", "Initiatives page loads (admin)",
     "Logged in as zain.admin",
     "1. Click Initiatives in sidebar",
     "Initiatives view renders without errors.",
     "P1"),
    ("Initiatives", "HL-INI-02", "Initiatives page loads (HOD)",
     "Logged in as hod.test",
     "1. Click Initiatives in sidebar",
     "Initiatives view renders without errors.",
     "P1"),
    ("Initiatives", "HL-INI-03", "Create an initiative",
     "On Initiatives page",
     "1. Click New Initiative\n2. Enter title, description, status\n3. Save",
     "Initiative appears in list with Active status.",
     "P2"),
]

# ── ADDITIONAL: AI Features ───────────────────────────────────────────────────

AI_CASES = [
    ("AI Suggest Standards", "AI-SUG-01", "AI Suggest Standards — returns suggestions",
     "Unit has an essential question. DeepSeek API key configured.",
     "1. Open a unit in UnitPlanView\n2. Open Map Standards section\n3. Click AI Suggest",
     "3–6 standards suggested based on unit theme. Suggestions are relevant to the essential question.",
     "P1"),
    ("AI Suggest Standards", "AI-SUG-02", "AI Suggest Standards — prioritises uncovered",
     "Some standards already mapped in other units",
     "1. Click AI Suggest on a unit",
     "Suggestions favour standards not yet mapped in the LTP.",
     "P2"),
    ("AI Suggest Standards", "AI-SUG-03", "AI Suggest Standards — no API key",
     "DEEPSEEK_API_KEY env var removed",
     "1. Click AI Suggest",
     "Error message shown. Does not crash the app.",
     "P2"),

    ("AI Fill Gaps", "AI-GAP-01", "AI Fill Gaps distributes unmapped standards",
     "LTP has multiple units, some standards unmapped",
     "1. Open LTP detail view\n2. Click AI Fill Gaps",
     "Unmapped standards distributed across existing units. No unit gets an unreasonable number.",
     "P1"),
    ("AI Fill Gaps", "AI-GAP-02", "AI Fill Gaps — no unmapped standards",
     "All standards already mapped",
     "1. Click AI Fill Gaps on a complete LTP",
     "Message shown: no gaps to fill. No changes made.",
     "P2"),

    ("AI Draft Full Year", "AI-DFT-01", "AI Draft Full Year generates units",
     "Empty LTP exists, DeepSeek key configured",
     "1. Open LTP with no units\n2. Click AI Draft Full Year\n3. Set unit count\n4. Confirm",
     "Complete set of units created with titles, essential questions, and standards distributed across terms.",
     "P1"),
    ("AI Draft Full Year", "AI-DFT-02", "AI Draft Full Year — streaming progress shown",
     "Triggering a draft",
     "1. Click AI Draft Full Year and watch UI",
     "Progress indicator shown while drafting. 'Drafting X units...' label visible.",
     "P2"),
    ("AI Draft Full Year", "AI-DFT-03", "AI Draft Full Year — cannot draft over existing units",
     "LTP has existing units",
     "1. Attempt to draft on LTP with units",
     "Warning or confirmation required. Existing units not silently overwritten.",
     "P1"),

    ("AI Draft Unit Content", "AI-UNT-01", "AI draft unit content fills rich fields",
     "Unit open in UnitPlanView, essential question set",
     "1. Click AI Draft (unit content)\n2. Wait for response",
     "Rich fields populated: learning outcomes, success criteria, anchor texts, vocabulary, lesson sequence.",
     "P1"),
    ("AI Draft Unit Content", "AI-UNT-02", "AI draft does not overwrite manually edited fields",
     "Some fields manually filled",
     "1. Manually edit Learning Outcomes\n2. Trigger AI draft",
     "Manually edited fields preserved. Only empty fields populated by AI.",
     "P2"),
]

# ── ADDITIONAL: PPT Generation ────────────────────────────────────────────────

PPT_CASES = [
    ("PPT Generation", "PPT-01", "Generate PPT from My Class view",
     "Teacher logged in, class has students and LTP data",
     "1. Navigate to My Class view\n2. Click Generate PPT / export option",
     "PPT generation sheet opens. Options for customisation visible.",
     "P1"),
    ("PPT Generation", "PPT-02", "PPT generation completes and downloads",
     "PPT generation sheet open",
     "1. Select options\n2. Click Generate\n3. Wait for completion",
     "PPTX file downloads to browser. File is valid and opens in PowerPoint/Keynote.",
     "P1"),
    ("PPT Generation", "PPT-03", "PPT generation with no units",
     "Class exists but LTP has no units",
     "1. Attempt PPT generation",
     "Error or empty-state message. Does not generate a corrupt file.",
     "P2"),
    ("PPT Generation", "PPT-04", "PPT generation — API error handling",
     "API call fails (e.g. timeout)",
     "1. Simulate API failure\n2. Trigger generation",
     "Error displayed in UI. Loading state clears. User can retry.",
     "P2"),
    ("PPT Generation", "PPT-05", "PPT requires teacher auth",
     "Unauthenticated request",
     "1. POST /api/generate-ppt without auth token",
     "Returns 401 Unauthorized.",
     "P1"),
]

# ── ADDITIONAL: Notifications ─────────────────────────────────────────────────

NOTIFY_CASES = [
    ("Notification Bell", "NOT-01", "Notification bell shows badge count",
     "Overdue units or pending reviews exist",
     "1. Log in as teacher with overdue items",
     "Bell icon shows numeric badge with count.",
     "P1"),
    ("Notification Bell", "NOT-02", "Click bell opens notification panel",
     "Notifications exist",
     "1. Click bell icon",
     "Notification panel slides open. Items listed with severity indicators.",
     "P1"),
    ("Notification Bell", "NOT-03", "Notification navigates to relevant view",
     "Panel open with notifications",
     "1. Click a notification item",
     "Panel closes. App navigates to the relevant view (e.g. HOD Review, LTP).",
     "P2"),
    ("Notification Bell", "NOT-04", "Mark all clear",
     "Notifications panel open",
     "1. Click All Clear",
     "Panel empties. Badge count clears.",
     "P2"),
    ("Notification Bell", "NOT-05", "HOD sees pending review notifications",
     "Teacher has submitted a unit",
     "1. Log in as HOD\n2. Check bell",
     "Notification for pending unit review visible.",
     "P1"),

    ("Email Digest", "NOT-06", "Send digest endpoint requires auth",
     "No auth token",
     "1. POST /api/notify/send-digest without token",
     "Returns 401 Unauthorized.",
     "P1"),
    ("Email Digest", "NOT-07", "Cron endpoint is protected",
     "No auth header",
     "1. GET /api/notify/cron without cron secret",
     "Returns 401. Cron secret required.",
     "P1"),
    ("Email Digest", "NOT-08", "Email digest sends to users with notification_email set",
     "User has notification_email configured",
     "1. Trigger /api/notify/send-digest with valid cron secret\n2. Check email inbox",
     "Digest email received with summary of overdue/pending items.",
     "P2"),
]

# ── ADDITIONAL: Delete LTP & Units ────────────────────────────────────────────

DELETE_LTP_CASES = [
    ("Delete LTP & Units", "TC-DEL-01", "Delete a Draft LTP",
     "Teacher has a Draft LTP",
     "1. Open LTP list\n2. Delete a Draft LTP",
     "LTP and all its units removed. No longer appears in list.",
     "P1"),
    ("Delete LTP & Units", "TC-DEL-02", "Cannot delete Approved LTP without confirmation",
     "Approved LTP exists",
     "1. Attempt to delete an Approved LTP",
     "Confirmation required or delete not available for approved plans.",
     "P1"),
    ("Delete LTP & Units", "TC-DEL-03", "Delete a Draft unit",
     "LTP has a Draft unit",
     "1. Open LTP detail\n2. Delete a Draft unit",
     "Unit removed. LTP unit list updates.",
     "P1"),
    ("Delete LTP & Units", "TC-DEL-04", "Cannot delete Approved unit",
     "LTP has an Approved unit",
     "1. Attempt to delete an Approved unit",
     "Delete not available or requires HOD action (reopen first).",
     "P1"),
    ("Delete LTP & Units", "TC-DEL-05", "HOD can reopen an Approved unit",
     "Approved unit exists",
     "1. HOD opens approved unit\n2. Clicks Re-open",
     "Unit status returns to Draft. Teacher can edit again.",
     "P1"),
]

# ── Assemble ───────────────────────────────────────────────────────────────────

ADMIN_CASES_FULL   = ADMIN_CASES + ADMIN_DELETE_CASES
TEACHER_CASES_FULL = TEACHER_CASES + DELETE_LTP_CASES

all_roles = [
    ("Platform Admin",    "BFDBFE", PLATFORM_CASES),
    ("School Admin",      "DDD6FE", ADMIN_CASES_FULL),
    ("Teacher / HOD",     "BBF7D0", TEACHER_CASES_FULL),
    ("HOD Leadership",    "FDE68A", HOD_LEADERSHIP_CASES),
    ("AI Features",       "E9D5FF", AI_CASES),
    ("PPT & Notify",      "FBCFE8", PPT_CASES + NOTIFY_CASES),
]

make_summary(wb, all_roles)
make_sheet(wb, "Platform Admin",  "3B82F6", PLATFORM_CASES,     "Platform Admin")
make_sheet(wb, "School Admin",    "7C3AED", ADMIN_CASES_FULL,   "School Admin")
make_sheet(wb, "Teacher & HOD",   "16A34A", TEACHER_CASES_FULL, "Teacher / HOD")
make_sheet(wb, "HOD Leadership",  "D97706", HOD_LEADERSHIP_CASES, "HOD Leadership")
make_sheet(wb, "AI Features",     "9333EA", AI_CASES,           "AI Features")
make_sheet(wb, "PPT & Notify",    "DB2777", PPT_CASES + NOTIFY_CASES, "PPT & Notifications")

# Remove default empty sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

total = (len(PLATFORM_CASES) + len(ADMIN_CASES_FULL) + len(TEACHER_CASES_FULL) +
         len(HOD_LEADERSHIP_CASES) + len(AI_CASES) + len(PPT_CASES) + len(NOTIFY_CASES))

out = "/Users/zainglenn/Documents/projects/education/curriculum-tracker/.design/TestPlan-CurriculumTracker-2026.xlsx"
wb.save(out)
print(f"Saved → {out}")
print(f"Total cases: {total}")
print(f"  Platform Admin:  {len(PLATFORM_CASES)}")
print(f"  School Admin:    {len(ADMIN_CASES_FULL)}")
print(f"  Teacher/HOD:     {len(TEACHER_CASES_FULL)}")
print(f"  HOD Leadership:  {len(HOD_LEADERSHIP_CASES)}")
print(f"  AI Features:     {len(AI_CASES)}")
print(f"  PPT & Notify:    {len(PPT_CASES) + len(NOTIFY_CASES)}")
